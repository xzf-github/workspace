'''
假设某学校所有课程每学期允许多次考试，学生可随时参加考试，系统自动将每次成绩
添加到 Excel 文件（包含 3 列：姓名，课程，成绩）中，现期末要求统计所有学生每门课
程的最高成绩。编写程序，模拟生成若干同学的成绩并写入 Excel 文件，
其中学生姓名和课程名称均可重复，也就是允许出现同一门课程的多次成绩，
最后统计所有学生每门课程的最高成绩，并写入新的 Excel 文件。
使用 pip install openpyxl 命令安装扩展库 openpyxl。
'''
import random
from openpyxl import Workbook,load_workbook

#生成随机数据
def generateRandomInfo(filename):
    workbook=Workbook()
    worksheet=workbook.worksheets[0]
    worksheet.append(['姓名','课程','成绩'])
    #百家姓: 赵钱孙李
    first ='赵钱孙李'
    middle='晨兴泽亿明'
    last='枫鹏曦海慧'
    subjects=('语文','数学','英语')
    for i in range(100):
        line=[]
        name=random.choice(first)
        #按一定概率生成只有两个字的中文名字
        if random.randint(1,100)>35:
            name=name+random.choice(middle)
        name=name+random.choice(last)
        #依次生成姓名,课程名称和成绩
        line.append(name)
        #for i in range(3):
        #    line.append(subjects[i])
        line.append(random.choice(subjects))
        line.append(random.randint(0,100))
        worksheet.append(line)
    #保存数据,生成excel 2007格式的文件
    workbook.save(filename)

def getResult(oldfile,newfile):
    #用于存放结果数据的字典
    result=dict()
    workbook=load_workbook(oldfile)
    worksheet=workbook.worksheets[0]
    #遍历原始数据
    for row in worksheet.rows:
        if row[0].value=='姓名':
            continue
        #姓名,课程名称,本次成绩
        name,subject,grade=row[0].value,row[1].value,row[2].value
        #获取当前姓名对应的课程名称和成绩信息
        #如果result字典中不包含,则返回空字典
        t=result.get(name,{})
        #获取当前学生当前课程的成绩,若不存在,返回0
        f=t.get(subject,0)
        #只保留该学生该课程的最高成绩
        if grade >f:
            t[subject]=grade
            result[name]=t
    workbook1=Workbook()
    worksheet1=workbook1.worksheets[0]
    worksheet1.append(['姓名','课程','成绩'])
    #将result字典中的结果数据写入excel文件
    for name,t in result.items():
        for subject,grade in t.items():
            worksheet1.append([name,subject,grade])
    workbook1.save(newfile)

oldfile=r'test.xlsx'
newfile=r'result.xlsx'
generateRandomInfo(oldfile)
getResult(oldfile,newfile)